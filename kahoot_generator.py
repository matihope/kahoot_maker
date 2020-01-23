import openpyxl
import random

with open('input.txt', 'r') as f:
    file_read = f.read().split('\n')

langs = file_read[0].split(' ')
slowka_i_def = [x.split(" - ") for x in file_read[1:]]

file = 'kahoot_template.xlsx'
wb = openpyxl.load_workbook(file)
ws = wb.active

for row, phrase in enumerate(slowka_i_def, 9):
    q = random.randrange(2)
    a = 1 - q

    answers = ['C', 'D', 'E', 'F']
    random.shuffle(answers)

    ws['B' + str(row)] = f"Przetłumacz na {langs[a]}: {phrase[q]}"
    ws[answers[0] + str(row)] = phrase[a]
    ws[answers[1] + str(row)] = slowka_i_def[random.randrange(len(slowka_i_def))][a]
    ws[answers[2] + str(row)] = slowka_i_def[random.randrange(len(slowka_i_def))][a]
    ws[answers[3] + str(row)] = slowka_i_def[random.randrange(len(slowka_i_def))][a]
    ws['G' + str(row)] = 10
    ws['H' + str(row)] = ord(answers[0]) - ord('B')

wb.save('generated_kahoot.xlsx')
